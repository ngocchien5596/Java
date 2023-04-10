import openpyxl

def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheetName = workbook[sheetName]
    return (sheetName.max_row)

def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheetName = workbook[sheetName]
    return (sheetName.max_column)

def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheetName = workbook[sheetName]
    return sheetName.cell(row = rownum, column = columnno).value

def writeData(file, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row = rownum, column = columnno).value = data
    workbook.save(file)
